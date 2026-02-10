import pandas as pd
import argparse
import os
import sys
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_excel(writer, df, sheet_name):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    bold_font = Font(bold=True)
    
    # Format headers
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    # Set column widths
    worksheet.column_dimensions['A'].width = 30
    for col_idx in range(2, 6): # B to E
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = 15

    # Highlight top results
    metric_cols = ['Top-1', 'Top-5', 'F1 Score']
    time_col = 'Time (min)'
    
    # Get column indices (1-based for openpyxl)
    col_indices = {name: df.columns.get_loc(name) + 1 for name in df.columns}
    
    for col_name in metric_cols:
        if col_name in df.columns:
            col_idx = col_indices[col_name]
            series = pd.to_numeric(df[col_name], errors='coerce')
            max_val = series.max()
            
            for row_idx, val in enumerate(series):
                if pd.notna(val) and val == max_val:
                    cell = worksheet.cell(row=row_idx + 2, column=col_idx) # +2 because header is row 1
                    cell.font = bold_font

            # Number format
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '0.0000'

    # Highlight min time
    if time_col in df.columns:
        col_idx = col_indices[time_col]
        series = pd.to_numeric(df[time_col], errors='coerce')
        min_val = series.min()
        
        for row_idx, val in enumerate(series):
            if pd.notna(val) and val == min_val:
                cell = worksheet.cell(row=row_idx + 2, column=col_idx)
                cell.font = bold_font
        
        # Number format
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.number_format = '0.0000'

def generate_report(input_csv, output_xlsx):
    if not os.path.exists(input_csv):
        print(f"Error: Input file '{input_csv}' not found.")
        sys.exit(1)
        
    try:
        df = pd.read_csv(input_csv)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        sys.exit(1)
        
    # config mapping
    config_map = {
        'fisher_only': 'Fisher (Ensemble)',
        'fisher_disk': 'Fisher (DISK)',
        'fisher_aliked': 'Fisher (ALIKED)',
        'fisher_superpoint': 'Fisher (SuperPoint)',
        'global_only': 'Global Only',
        'global_fisher': 'Global + Fisher',
        'global_fisher_gv_power': 'Global + Fisher + GV (3-Tier)',
        'fisher_gv_power': 'Fisher + GV',
        'wildfusion_baseline': 'WildFusion Baseline',
        'global_gv': 'Global + GV'
    }
    
    # Filter for ok status
    df = df[df['status'] == 'ok'].copy()
    
    if df.empty:
        print("No successful runs found in CSV.")
        sys.exit(0)

    # Normalize dataset names
    df['dataset'] = df['dataset'].str.strip().str.lower()
    
    datasets = sorted(df['dataset'].unique())
    
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        for ds in datasets:
            print(f"Processing dataset: {ds}")
            ds_df = df[df['dataset'] == ds].copy()
            
            report_data = []
            
            for idx, row in ds_df.iterrows():
                cfg = row['config']
                method_name = config_map.get(cfg, cfg)
                
                try:
                    top1 = float(row['accuracy']) if pd.notna(row['accuracy']) and row['accuracy'] != '' else None
                    top5 = float(row['top5_accuracy']) if pd.notna(row['top5_accuracy']) and row['top5_accuracy'] != '' else None
                    f1 = float(row['f1_score']) if pd.notna(row['f1_score']) and row['f1_score'] != '' else None
                    time_min = float(row['runtime_minutes']) if pd.notna(row['runtime_minutes']) and row['runtime_minutes'] != '' else None
                except ValueError:
                    continue

                report_data.append({
                    'Method': method_name,
                    'Top-1': top1,
                    'Top-5': top5,
                    'F1 Score': f1,
                    'Time (min)': time_min
                })
                
            report_df = pd.DataFrame(report_data)
            
            if not report_df.empty:
                # Remove duplicates if any (same config run multiple times) - keep last
                report_df = report_df.drop_duplicates(subset=['Method'], keep='last')
                # Sort by Method name for consistency
                report_df = report_df.sort_values('Method')
                
                safe_sheet_name = ds[:30].replace('/', '_')
                format_excel(writer, report_df, safe_sheet_name)
            else:
                print(f"Skipping empty dataset {ds}")

    print(f"Report generated: {output_xlsx}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', default='evaluations/classifications/final_comparisons.csv')
    parser.add_argument('--output', default='evaluations/classifications/final_comparison_report.xlsx')
    args = parser.parse_args()
    
    generate_report(args.input, args.output)
