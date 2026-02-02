from __future__ import annotations

from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from scipy import stats


INPUT_XLSX_PATH = (
    Path(__file__).resolve().parents[1]
    / "evaluations"
    / "classification"
    / "classification_results_ALL_FIXED_SPLIT_with_z_score.xlsx"
)

OUTPUT_DIR = Path(__file__).resolve().parents[1] / "Output" / "classification_analysis"
OUTPUT_XLSX_PATH = OUTPUT_DIR / "classification_results_analysis.xlsx"
FIGURES_DIR = OUTPUT_DIR / "figures"

METRIC_COLUMNS = ["Accuracy", "Top-5 Accuracy", "F-1 Score", "Run Time (minutes)"]
NON_CONFIG_COLUMNS = ["Dataset", "Training Examples", "Num Classes", "Dataset Type"]

# Explicitly ignored for config identity (per request)
IGNORED_CONFIG_COLUMNS = {
    "GMM Components",
    "PCA Components",
    "Embedding Model",
    "Method",
    "GV Method",
    "Alpha (fv sim - gv)",
    "Geom. Candidates",
    "Min Inliers",
    "Inlier Threshold",
    "MAX GMM Descriptors (per image)",
}

PREFERRED_FACTOR_ORDER = [
    "Use Global Embedding",
    "Use Fisher",
    "Use GV",
    "MD Split Type",
    "MD Trained On",
    "MD Random Split",
    "Global Weight",
    "Fisher Weight",
    "Remove Background",
]


def _is_nan(value: Any) -> bool:
    return value is None or (isinstance(value, float) and np.isnan(value)) or pd.isna(value)


def _fmt_value(value: Any) -> str:
    if _is_nan(value):
        return "NA"
    if isinstance(value, (np.bool_, bool)):
        return "T" if bool(value) else "F"
    if isinstance(value, (np.integer, int)):
        return str(int(value))
    if isinstance(value, (np.floating, float)):
        v = float(value)
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return f"{v:.4g}"
    return str(value)


def _select_factor_columns(df: pd.DataFrame) -> list[str]:
    config_cols = [c for c in df.columns if c not in set(METRIC_COLUMNS + NON_CONFIG_COLUMNS)]
    config_cols = [c for c in config_cols if c not in IGNORED_CONFIG_COLUMNS]

    kept: list[str] = []
    for col in config_cols:
        non_nan = df[col].dropna()
        if non_nan.nunique() <= 1:
            continue
        kept.append(col)

    ordered: list[str] = []
    for c in PREFERRED_FACTOR_ORDER:
        if c in kept:
            ordered.append(c)
    ordered.extend(sorted([c for c in kept if c not in set(ordered)]))
    return ordered


def _config_desc_from_factors(row: pd.Series, factor_cols: list[str]) -> str:
    return " | ".join([f"{c}={_fmt_value(row.get(c))}" for c in factor_cols])


def _safe_wilcoxon_p(deltas: np.ndarray) -> float | None:
    deltas = deltas.astype(float)
    deltas = deltas[~np.isnan(deltas)]
    if len(deltas) < 5:
        return None
    if np.allclose(deltas, 0):
        return 1.0
    try:
        res = stats.wilcoxon(deltas, zero_method="wilcox", alternative="two-sided")
        return float(res.pvalue)
    except Exception:
        return None


def _pareto_flags(df: pd.DataFrame, *, x_col: str, y_col: str) -> pd.DataFrame:
    base = df.dropna(subset=[x_col, y_col]).copy()
    if base.empty:
        base["is_dominated"] = False
        base["is_pareto"] = False
        return base

    points = base[[x_col, y_col]].to_numpy(dtype=float)
    is_dominated = np.zeros(len(points), dtype=bool)
    for i in range(len(points)):
        if is_dominated[i]:
            continue
        x_i, y_i = points[i]
        for j in range(len(points)):
            if i == j:
                continue
            x_j, y_j = points[j]
            if (y_j >= y_i and x_j <= x_i) and (y_j > y_i or x_j < x_i):
                is_dominated[i] = True
                break

    base["is_dominated"] = is_dominated
    base["is_pareto"] = ~base["is_dominated"]
    return base


def _write_excel_tables(xlsx_path: Path, tables: dict[str, pd.DataFrame]) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet, table in tables.items():
            table.to_excel(writer, sheet_name=sheet, index=False)


def _plot_figures(
    *,
    fig_dir: Path,
    config_summary: pd.DataFrame,
    toggle_deltas: pd.DataFrame,
    dataset_best: pd.DataFrame,
) -> list[Path]:
    fig_dir.mkdir(parents=True, exist_ok=True)
    sns.set_theme(style="whitegrid")

    images: list[Path] = []

    top = config_summary.sort_values(["f1_mean", "acc_mean"], ascending=[False, False]).head(12).copy()
    if not top.empty:
        plt.figure(figsize=(14, 6))
        ax = sns.barplot(data=top, x="config_id", y="f1_mean", color="#4C78A8")
        ax.set_title("Top configurations by mean best F1 across datasets")
        ax.set_xlabel("config_id")
        ax.set_ylabel("mean best F1")
        ax.tick_params(axis="x", rotation=45)
        plt.tight_layout()
        out = fig_dir / "top_configs_mean_f1.png"
        plt.savefig(out, dpi=160)
        plt.close()
        images.append(out)

    runtime_base = config_summary.dropna(subset=["rt_mean", "f1_mean"]).copy()
    if not runtime_base.empty:
        pareto = _pareto_flags(runtime_base, x_col="rt_mean", y_col="f1_mean")
        plt.figure(figsize=(10, 6))
        ax = sns.scatterplot(
            data=pareto,
            x="rt_mean",
            y="f1_mean",
            hue="is_pareto",
            palette={True: "#F58518", False: "#B0B0B0"},
            alpha=0.9,
        )
        ax.set_title("Runtime vs mean best F1 (Pareto frontier highlighted)")
        ax.set_xlabel("mean runtime (minutes)")
        ax.set_ylabel("mean best F1")
        ax.legend(title="Pareto")
        plt.tight_layout()
        out = fig_dir / "runtime_vs_f1_pareto.png"
        plt.savefig(out, dpi=160)
        plt.close()
        images.append(out)

    if not toggle_deltas.empty and "delta_f1" in toggle_deltas.columns:
        plt.figure(figsize=(10, 5))
        ax = sns.violinplot(data=toggle_deltas, x="toggle", y="delta_f1", inner="quartile", color="#72B7B2")
        ax.axhline(0, color="black", linewidth=1, alpha=0.6)
        ax.set_title("Paired per-dataset delta: best-on minus best-off (F1)")
        ax.set_xlabel("toggle")
        ax.set_ylabel("ΔF1")
        ax.tick_params(axis="x", rotation=20)
        plt.tight_layout()
        out = fig_dir / "toggle_deltas_f1.png"
        plt.savefig(out, dpi=160)
        plt.close()
        images.append(out)

    if not dataset_best.empty:
        ds = dataset_best.sort_values("best_f1", ascending=True).copy()
        plt.figure(figsize=(12, 10))
        ax = sns.barplot(data=ds, y="Dataset", x="best_f1", color="#54A24B")
        ax.set_title("Best achievable F1 per dataset (after filtering)")
        ax.set_xlabel("best F1")
        ax.set_ylabel("Dataset")
        plt.tight_layout()
        out = fig_dir / "dataset_best_f1.png"
        plt.savefig(out, dpi=160)
        plt.close()
        images.append(out)

    return images


def _embed_images_in_workbook(xlsx_path: Path, image_paths: list[Path]) -> None:
    wb = load_workbook(xlsx_path)
    if "Charts" in wb.sheetnames:
        ws = wb["Charts"]
        wb.remove(ws)
    ws = wb.create_sheet("Charts")
    ws["A1"] = "Charts (generated)"

    row = 3
    for img_path in image_paths:
        if not img_path.exists():
            continue
        ws[f"A{row}"] = img_path.name
        row += 1
        img = XLImage(str(img_path))
        img.anchor = f"A{row}"
        ws.add_image(img)
        row += 28

    wb.save(xlsx_path)


def main() -> None:
    if not INPUT_XLSX_PATH.exists():
        raise FileNotFoundError(f"Missing input file: {INPUT_XLSX_PATH}")

    df_raw = pd.read_excel(INPUT_XLSX_PATH, sheet_name="Sheet1")
    missing_cols = [c for c in ["Dataset", *METRIC_COLUMNS] if c not in df_raw.columns]
    if missing_cols:
        raise ValueError(f"Input is missing required columns: {missing_cols}")

    # Filter: ignore datasets that only have 1 entry in the table
    dataset_sizes = df_raw.groupby("Dataset").size()
    kept_datasets = set(dataset_sizes[dataset_sizes >= 2].index)
    dropped_datasets = sorted(set(df_raw["Dataset"].unique()) - kept_datasets)
    df = df_raw[df_raw["Dataset"].isin(kept_datasets)].copy()

    # Build simplified configuration identity (ignores GMM/PCA/Embedding Model/Method/etc)
    factor_cols = _select_factor_columns(df)
    df["config_desc"] = df.apply(lambda r: _config_desc_from_factors(r, factor_cols), axis=1)
    unique_descs = sorted(df["config_desc"].unique())
    desc_to_id = {desc: f"C{idx:02d}" for idx, desc in enumerate(unique_descs, start=1)}
    df["config_id"] = df["config_desc"].map(desc_to_id)

    configs = (
        df[["config_id", "config_desc", *factor_cols]]
        .drop_duplicates("config_id")
        .sort_values("config_id")
        .reset_index(drop=True)
    )

    # Keep output tables clean: exclude ignored columns from the row table.
    row_cols = [
        *[c for c in NON_CONFIG_COLUMNS if c in df.columns],
        "config_id",
        "config_desc",
        *factor_cols,
        *[c for c in METRIC_COLUMNS if c in df.columns],
    ]
    rows = df[row_cols].sort_values(["Dataset", "config_id"]).reset_index(drop=True)

    # Collapse by dataset + simplified config_id (metrics: best performance within the simplified config)
    agg: dict[str, Any] = {
        "config_desc": "first",
        "Training Examples": "first",
        "Num Classes": "first",
        "Dataset Type": "first",
        "Accuracy": "max",
        "Top-5 Accuracy": "max",
        "F-1 Score": "max",
        "Run Time (minutes)": "min",
    }
    for c in factor_cols:
        agg[c] = "first"
    rows_collapsed = (
        df.groupby(["Dataset", "config_id"], dropna=False)
        .agg(agg)
        .reset_index()
        .rename(
            columns={
                "Accuracy": "best_acc",
                "Top-5 Accuracy": "best_top5",
                "F-1 Score": "best_f1",
                "Run Time (minutes)": "best_runtime_min",
            }
        )
        .sort_values(["Dataset", "config_id"])
        .reset_index(drop=True)
    )

    # Per-dataset winners (on collapsed table)
    winners: list[dict[str, Any]] = []
    for dataset, g in rows_collapsed.groupby("Dataset"):
        g_f1 = g.sort_values("best_f1", ascending=False)
        g_acc = g.sort_values("best_acc", ascending=False)

        best_f1 = g_f1.iloc[0]
        best_acc = g_acc.iloc[0]

        f1_margin = float(best_f1["best_f1"] - g_f1.iloc[1]["best_f1"]) if len(g_f1) >= 2 else np.nan
        acc_margin = float(best_acc["best_acc"] - g_acc.iloc[1]["best_acc"]) if len(g_acc) >= 2 else np.nan

        winners.append(
            {
                "Dataset": dataset,
                "Training Examples": best_f1.get("Training Examples"),
                "Num Classes": best_f1.get("Num Classes"),
                "n_configs": int(g["config_id"].nunique()),
                "best_f1": float(best_f1["best_f1"]),
                "best_f1_margin_vs_2nd": f1_margin,
                "best_f1_config_id": best_f1["config_id"],
                "best_f1_config_desc": best_f1["config_desc"],
                "best_acc": float(best_acc["best_acc"]),
                "best_acc_margin_vs_2nd": acc_margin,
                "best_acc_config_id": best_acc["config_id"],
                "best_acc_config_desc": best_acc["config_desc"],
                "best_runtime_min": float(best_f1["best_runtime_min"])
                if not _is_nan(best_f1["best_runtime_min"])
                else np.nan,
            }
        )
    dataset_winners = pd.DataFrame(winners).sort_values("best_f1", ascending=False).reset_index(drop=True)

    # Configuration summary across datasets (using collapsed maxima per dataset)
    summary = (
        rows_collapsed.groupby("config_id", dropna=False)
        .agg(
            datasets=("Dataset", "nunique"),
            rows=("Dataset", "size"),
            f1_mean=("best_f1", "mean"),
            f1_median=("best_f1", "median"),
            f1_std=("best_f1", "std"),
            acc_mean=("best_acc", "mean"),
            acc_median=("best_acc", "median"),
            acc_std=("best_acc", "std"),
            top5_mean=("best_top5", "mean"),
            rt_mean=("best_runtime_min", "mean"),
            rt_median=("best_runtime_min", "median"),
        )
        .reset_index()
        .merge(configs, on="config_id", how="left")
    )

    # Win counts
    best_f1_idx = rows_collapsed.groupby("Dataset")["best_f1"].idxmax()
    best_acc_idx = rows_collapsed.groupby("Dataset")["best_acc"].idxmax()
    f1_wins = rows_collapsed.loc[best_f1_idx, "config_id"].value_counts()
    acc_wins = rows_collapsed.loc[best_acc_idx, "config_id"].value_counts()
    summary["f1_wins"] = summary["config_id"].map(f1_wins).fillna(0).astype(int)
    summary["acc_wins"] = summary["config_id"].map(acc_wins).fillna(0).astype(int)

    config_summary = summary.sort_values(["f1_mean", "acc_mean", "rt_mean"], ascending=[False, False, True]).reset_index(
        drop=True
    )

    # Toggle deltas (paired per dataset: best-on minus best-off)
    toggles = [t for t in ["Use Global Embedding", "Use Fisher", "Use GV"] if t in rows_collapsed.columns]
    delta_rows: list[dict[str, Any]] = []
    for toggle in toggles:
        for dataset, g in rows_collapsed.groupby("Dataset"):
            if g[toggle].dropna().nunique() < 2:
                continue
            on = g[g[toggle] == True]  # noqa: E712
            off = g[g[toggle] == False]  # noqa: E712
            if on.empty or off.empty:
                continue

            on_best_f1 = on.sort_values("best_f1", ascending=False).iloc[0]
            off_best_f1 = off.sort_values("best_f1", ascending=False).iloc[0]
            on_best_acc = on.sort_values("best_acc", ascending=False).iloc[0]
            off_best_acc = off.sort_values("best_acc", ascending=False).iloc[0]

            delta_rows.append(
                {
                    "Dataset": dataset,
                    "toggle": toggle,
                    "on_best_f1_config_id": on_best_f1["config_id"],
                    "off_best_f1_config_id": off_best_f1["config_id"],
                    "on_best_f1": float(on_best_f1["best_f1"]),
                    "off_best_f1": float(off_best_f1["best_f1"]),
                    "delta_f1": float(on_best_f1["best_f1"]) - float(off_best_f1["best_f1"]),
                    "on_best_acc_config_id": on_best_acc["config_id"],
                    "off_best_acc_config_id": off_best_acc["config_id"],
                    "on_best_acc": float(on_best_acc["best_acc"]),
                    "off_best_acc": float(off_best_acc["best_acc"]),
                    "delta_acc": float(on_best_acc["best_acc"]) - float(off_best_acc["best_acc"]),
                }
            )
    toggle_deltas = pd.DataFrame(delta_rows).sort_values(["toggle", "Dataset"]).reset_index(drop=True)

    toggle_summary_rows: list[dict[str, Any]] = []
    for toggle, g in toggle_deltas.groupby("toggle"):
        for metric in ["delta_f1", "delta_acc"]:
            arr = g[metric].to_numpy(dtype=float)
            n = int(np.sum(~np.isnan(arr)))
            pos = int(np.sum(arr[~np.isnan(arr)] > 0))
            neg = int(np.sum(arr[~np.isnan(arr)] < 0))
            zero = n - pos - neg
            toggle_summary_rows.append(
                {
                    "toggle": toggle,
                    "metric": metric,
                    "n": n,
                    "mean": float(np.nanmean(arr)) if n else np.nan,
                    "median": float(np.nanmedian(arr)) if n else np.nan,
                    "pos": pos,
                    "neg": neg,
                    "zero": zero,
                    "wilcoxon_p": _safe_wilcoxon_p(arr),
                }
            )
    toggle_summary = pd.DataFrame(toggle_summary_rows).sort_values(["metric", "toggle"]).reset_index(drop=True)

    # Pareto: runtime vs mean best F1
    pareto = _pareto_flags(config_summary, x_col="rt_mean", y_col="f1_mean")[
        ["config_id", "datasets", "f1_mean", "acc_mean", "rt_mean", "is_pareto", "is_dominated", "config_desc", *factor_cols]
    ].sort_values(["is_pareto", "f1_mean", "rt_mean"], ascending=[False, False, True]).reset_index(drop=True)

    # Correlations (dataset-level)
    cor_rows: list[dict[str, Any]] = []
    for xcol in ["Training Examples", "Num Classes"]:
        if xcol not in dataset_winners.columns:
            continue
        x = dataset_winners[xcol].astype(float)
        y_f1 = dataset_winners["best_f1"].astype(float)
        y_acc = dataset_winners["best_acc"].astype(float)
        spear_f1 = stats.spearmanr(x, y_f1, nan_policy="omit")
        spear_acc = stats.spearmanr(x, y_acc, nan_policy="omit")
        cor_rows.append(
            {"x": xcol, "y": "best_f1", "method": "spearman", "value": spear_f1.correlation, "p_value": spear_f1.pvalue}
        )
        cor_rows.append(
            {"x": xcol, "y": "best_acc", "method": "spearman", "value": spear_acc.correlation, "p_value": spear_acc.pvalue}
        )
    correlations = pd.DataFrame(cor_rows)

    readme = pd.DataFrame(
        [
            {"Field": "Input XLSX", "Value": str(INPUT_XLSX_PATH)},
            {"Field": "Output XLSX", "Value": str(OUTPUT_XLSX_PATH)},
            {"Field": "Kept datasets (>=2 rows)", "Value": int(len(kept_datasets))},
            {"Field": "Dropped datasets (<2 rows)", "Value": int(len(dropped_datasets))},
            {"Field": "Dropped datasets list", "Value": ", ".join(dropped_datasets)},
            {"Field": "Ignored config columns", "Value": ", ".join(sorted(IGNORED_CONFIG_COLUMNS))},
            {"Field": "Factor columns used for config_id", "Value": ", ".join(factor_cols)},
            {
                "Field": "Rows_Collapsed meaning",
                "Value": "Per Dataset + config_id, metrics are MAX(Accuracy/F1/Top5) and MIN(Runtime) within that simplified config.",
            },
        ]
    )

    figures = _plot_figures(
        fig_dir=FIGURES_DIR,
        config_summary=config_summary,
        toggle_deltas=toggle_deltas,
        dataset_best=dataset_winners[["Dataset", "best_f1"]],
    )

    tables = {
        "README": readme,
        "Rows": rows,
        "Rows_Collapsed": rows_collapsed,
        "Configs": configs,
        "Config_Summary": config_summary,
        "Dataset_Winners": dataset_winners,
        "Toggle_Deltas": toggle_deltas,
        "Toggle_Summary": toggle_summary,
        "Pareto": pareto,
        "Correlations": correlations,
    }
    _write_excel_tables(OUTPUT_XLSX_PATH, tables)
    _embed_images_in_workbook(OUTPUT_XLSX_PATH, figures)

    print(f"Wrote XLSX: {OUTPUT_XLSX_PATH}")
    print(f"Wrote figures: {FIGURES_DIR}")


if __name__ == "__main__":
    main()
